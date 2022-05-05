from tkinter import CENTER
from openpyxl import load_workbook  
from openpyxl.styles import Font, colors, alignment, Color
from openpyxl.styles import NamedStyle

# How to import path excel file
wb = load_workbook(r'C:\Users\LENOVO\Downloads\Analysis\amira.xlsx')  
# identify Current sheet in the workbook
sh1 = wb['Test'] 
# identify cell to enter your input and choose style
cell = sh1['f1']
cell.value = 'welcome python'
cell.font = Font(color= colors.BLUE, size = 20,italic= True)
sh1.merge_cells("f1:g1")
wb.save('amira.xlsx')

#print(wb.active.title)