# first import load_workbook if you have alread an existing excel file
from openpyxl import load_workbook

# creat var and use load_workbook and write the file path for your excel file
workbook = load_workbook(r'C:\Users\LENOVO\Downloads\Analysis\NoofClick.xlsx')
sheet = workbook.active

#start to work
sheet["a1"].value = "Month"

sheet["b1"].value = "Link Clicks"
print(sheet["a1"].value)
print(sheet["b1"].value)

workbook.save(r"C:\Users\LENOVO\Downloads\Analysis\NoofClick.xlsx")
#noted the code will not run successfull esle you close the Excel file first and after run it back to your excel file to check the result 
